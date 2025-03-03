
from django.shortcuts import render, redirect
from .forms import ContactForm
from django.shortcuts import render, redirect, get_object_or_404
from .models import Contact  # Your Contact model
from .forms import ContactForm  # Your ContactForm
from .forms import LoginForm
from .models import Authenticate as authenticate
from django.contrib.auth.decorators import login_required


@login_required
def contact_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()  # Save data to the database
            return render(request, 'thank_you.html', {'name': form.cleaned_data['name']})
    else:
        form = ContactForm()
    return render(request, 'contact_form.html', {'form': form})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Contact
import json

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Contact  # Make sure your model is correctly imported

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Contact  # Ensure your model is correctly imported

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from myapp.models import Contact, DataRecordBVS # Ensure correct imports

@csrf_exempt  # Remove this if CSRF protection is needed
def check_BVS_Device(request):
    BVS_Device = request.GET.get('BVS_Device')
    
    if not BVS_Device:
        return JsonResponse({'error': 'BVS_Device parameter is required'}, status=400)

    contact = Contact.objects.filter(BVS_Device=BVS_Device).first()  # Get the first match if exists

    if contact:
        return JsonResponse({'exists': True, 'retailer_id': contact.Retailer_ID, 'category': contact.Category})

    # Step 2: Check if Device_ID exists in DataRecord table
    data_record = DataRecordBVS.objects.filter(Device_ID=BVS_Device).first()

    if data_record:
        franchise_id = data_record.Franchise_ID  # Fixed field name
        print(f"Fetched Franchise_ID from DB: {franchise_id}")  # Debugging

        # Step 3: Check if Franchise_ID matches the logged-in username
        username = request.user.username if request.user.is_authenticated else None  # Fetching username

        if franchise_id != username:
            return JsonResponse({'exists': True, 'error': 'You are not authorized to use this BVS Device.'}, status=403)

    return JsonResponse({'exists': False})  # Return false if no record is found

    return JsonResponse({'exists': False})
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
import pytz
from django.utils.timezone import now
from .models import Contact, edit_contact  # Adjust import based on your project structure

@csrf_exempt  # Remove if CSRF protection is enabled
def update_BVS_Device(request):
    if request.method == "POST":
        data = json.loads(request.body)
        BVS_Device = data.get("BVS_Device")

        # Find the existing contact with this BVS device
        contact = Contact.objects.filter(BVS_Device=BVS_Device).first()
        if contact:
            # Convert timestamp to Pakistan Standard Time (PST)
            pakistan_tz = pytz.timezone('Asia/Karachi')
            pakistan_time = now().astimezone(pakistan_tz)

            # Move previous record to edit_contact before updating the Contact model
            edit_contact.objects.create(
                Retailer_ID=contact.Retailer_ID,
                Franchise_ID=contact.Franchise_ID,
                Retailer_Number=contact.Retailer_Number,
                DSO_Name=contact.DSO_Name,
                CNIC=contact.CNIC,
                BVS_Device=contact.BVS_Device,  # Store previous BVS device
                Location=contact.Location,
                username=request.user.username if request.user.is_authenticated else "Unknown",
                Category=contact.Category,
                Other_Contact_Number=contact.Other_Contact_Number,
                Date_of_Joining=contact.Date_of_Joining,
                Date_of_Resignation=contact.Date_of_Resignation,  # Maintain existing resignation date
                Date_of_Data_Entry=pakistan_time
            )

            # Set BVS_Device to None for the existing record
            contact.BVS_Device = None
            contact.save()

            return JsonResponse({"success": True})

    return JsonResponse({"success": False})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from myapp.models import Contact, DataRecord  # Ensure these models are imported

@csrf_exempt  
@login_required
def check_retailer_id(request):
    retailer_id = request.GET.get('retailer_id')
    print(retailer_id)
    username = request.user.username  # Get logged-in user's username

    print(f"Logged-in Username: {username}")  # Debugging

    if retailer_id:
        # ✅ Step 1: Check if Retailer ID exists in Contact table
        if Contact.objects.filter(Retailer_ID=retailer_id).exists():
            return JsonResponse({'exists': True, 'error': 'Retailer ID already exists.'}, status=400)

        try:
            # ✅ Step 2: Check if Retailer ID exists in DataRecord table
            data_record = DataRecord.objects.get(retailer_id=retailer_id)  # Fixed field name
            franchise_id = data_record.second_parent_id  # Fixed field name

            print(f"Fetched Franchise_ID from DB: {franchise_id}")  # Debugging

            # ✅ Step 3: Check if Franchise_ID matches the logged-in username
            if franchise_id != username:
                return JsonResponse({'exists': True, 'error': 'You are not authorized to use this Retailer ID.'}, status=403)

            # ✅ Step 4: Allow submission if Retailer ID exists but Franchise matches
            return JsonResponse({'exists': False, 'authorized': True})  # ✅ Changed from 400 error to success

        except DataRecord.DoesNotExist:
            # ✅ Step 5: If Retailer ID does not exist in both tables, allow submission
            return JsonResponse({'exists': False, 'authorized': True})

    return JsonResponse({'exists': False, 'authorized': False})


from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import now
from .models import ContactChangeLog
import pytz
from django.utils.timezone import now

# @login_required
# def contact_edit_view(request, pk):
#     contact = get_object_or_404(Contact, pk=pk)

#     if request.method == 'POST':
#         form = ContactForm(request.POST, instance=contact)

#         if form.is_valid():
#             # Fetch the original instance from the database again to ensure consistency
#             original_contact = Contact.objects.get(pk=pk)
            
#             # Track changes dynamically
#             changes = {}

#             # Get the dynamically changed fields from the form
#             for field in form.changed_data:
#                 old_value = getattr(original_contact, field)  # Old value from the database
#                 new_value = form.cleaned_data[field]  # New value from the form

#                 # Add the change to the dictionary
#                 changes[field] = {
#                     "old": old_value,
#                     "new": new_value,
#                 }

#             # Save the updated contact
#             form.save()

#             # Convert the timestamp to Pakistan Standard Time (PST)
#             pakistan_tz = pytz.timezone('Asia/Karachi')
#             timestamp = now().astimezone(pakistan_tz)

#             # Log changes dynamically
#             for field, change in changes.items():
#                 ContactChangeLog.objects.create(
#                     username=request.user.username,
#                     contact=contact,
#                     field_name=field,
#                     old_value=change["old"],
#                     new_value=change["new"],
#                     timestamp=timestamp,
#                 )

#             return redirect('contact_edit', pk=pk)

#     else:
#         form = ContactForm(instance=contact)

#     # Pass the original field values to the template
#     initial_values = {field.name: getattr(contact, field.name) for field in form}

#     return render(request, 'edit_contact.html', {'form': form, 'contact': contact, 'initial_values': initial_values})
from django.utils.timezone import now
import pytz
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Contact, ContactChangeLog, edit_contact  # Ensure EditContact model is defined
from .forms import ContactForm

@login_required
def contact_edit_view(request, pk):
    contact = get_object_or_404(Contact, pk=pk)
    user = request.user  # Get logged-in user

    if request.method == 'POST':
        form = ContactForm(request.POST, instance=contact)

        if form.is_valid():
            # Fetch the original instance before updating
            original_contact = Contact.objects.get(pk=pk)

            # Convert timestamp to Pakistan Standard Time (PST)
            pakistan_tz = pytz.timezone('Asia/Karachi')
            pakistan_time = now().astimezone(pakistan_tz)

            # Move record to edit_contact table before updating
            edit_contact.objects.create(
                Retailer_ID=original_contact.Retailer_ID,
                Franchise_ID=original_contact.Franchise_ID,
                Retailer_Number=original_contact.Retailer_Number,
                DSO_Name=original_contact.DSO_Name,
                CNIC=original_contact.CNIC,
                BVS_Device=original_contact.BVS_Device,
                Location=original_contact.Location,
                username=user.username,  # Store the username of the logged-in user
                Category=original_contact.Category,
                Other_Contact_Number=original_contact.Other_Contact_Number,
                Date_of_Joining=original_contact.Date_of_Joining,
                Date_of_Resignation=original_contact.Date_of_Resignation,  # Set resignation date to today
                Date_of_Data_Entry=pakistan_time
            )

            # Track changes dynamically
            changes = {}

            for field in form.changed_data:
                old_value = getattr(original_contact, field)
                new_value = form.cleaned_data[field]

                changes[field] = {"old": old_value, "new": new_value}

            # Save the updated contact
            form.save()

            # Log changes dynamically
            for field, change in changes.items():
                ContactChangeLog.objects.create(
                    username=user.username,
                    contact=contact,
                    field_name=field,
                    old_value=change["old"],
                    new_value=change["new"],
                    timestamp=pakistan_time,
                )

            return redirect('contact_edit', pk=pk)

    else:
        form = ContactForm(instance=contact)

    initial_values = {field.name: getattr(contact, field.name) for field in form}

    return render(request, 'edit_contact.html', {'form': form, 'contact': contact, 'initial_values': initial_values})

from django.shortcuts import redirect


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Contact
from .forms import ContactForm

@login_required
def contact_list_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('contact_list')

    else:
        form = ContactForm()

    # Filter contacts where Franchise_ID matches the logged-in user's username
    contacts = Contact.objects.filter(Category='Retailer', Franchise_ID=request.user.username)
    
    return render(request, 'contact_list.html', {'form': form, 'contacts': contacts})


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Contact
from .forms import ContactForm

@login_required
def DSO_list_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()  # Save data to the database
            return redirect('DSO_List')  # Redirect after saving

    else:
        form = ContactForm()

    # Fetch contacts where Category is 'DSO' and Date_of_Resignation is NULL
    contacts = Contact.objects.filter(Category='DSO', Franchise_ID=request.user.username,Date_of_Resignation__isnull=True)

    return render(request, 'DSO_List.html', {'form': form, 'contacts': contacts})


@login_required
def RSO_list_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()  # Save data to the database
            # Redirect to the same page after saving the form data
            return redirect('RSO_list')  # Replace 'contact_list' with your URL name for the contact list view
         
    else:
        form = ContactForm()

    contacts = Contact.objects.filter(Category='RSO',Franchise_ID=request.user.username) # Fetch all records from the database
    return render(request, 'RSO_list.html', {'form': form, 'contacts': contacts})   
@login_required
def WIC_list_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()  # Save data to the database
            # Redirect to the same page after saving the form data
            return redirect('WIC_list')  # Replace 'contact_list' with your URL name for the contact list view
         
    else:
        form = ContactForm()

    contacts = Contact.objects.filter(Category='WIC',Franchise_ID=request.user.username) # Fetch all records from the database
    return render(request, 'WIC_list.html', {'form': form, 'contacts': contacts})   
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from .forms import LoginForm
from django.contrib.auth.models import User

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from .forms import LoginForm

def login_view(request):
    form = LoginForm(request.POST or None)
    msg = None

    if request.method == "POST" and form.is_valid():
        username = form.cleaned_data.get("username")
        password = form.cleaned_data.get("password")
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("/contact/list")  # Redirect to home or dashboard page
        else:
            msg = "Invalid credentials"
    elif request.method == "POST":
        msg = "Error validating the form"

    return render(request, "Login.html", {"form": form, "msg": msg})



import io
import csv
from django.http import HttpResponse
from django.shortcuts import render
from .forms import FileUploadForm
from .models import DataRecord




from django.http import JsonResponse
from .models import DataRecord

def get_retailer_number(request):
    retailer_id = request.GET.get('retailer_id')
    if retailer_id:
        try:
            data_record = DataRecord.objects.get(retailer_id=retailer_id)
           
            return JsonResponse({'retailer_number': data_record.retailer_msisdn,'Franchise_ID': data_record.second_parent_id,
            'Region': data_record.second_parent_region})
        except DataRecord.DoesNotExist:
            return JsonResponse({'retailer_number': None})
    return JsonResponse({'retailer_number': None})

import io
import csv
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import load_workbook
from .forms import FileUploadForm, BVSUploadForm, HeirarchyUploadForm
from .models import DataRecord, DataRecordBVS, Heirarchy

def upload_file(request):
    form1 = FileUploadForm()
    form2 = BVSUploadForm()
    form3 = HeirarchyUploadForm()

    if request.method == 'POST':
        if 'upload_csv' in request.POST:
            form1 = FileUploadForm(request.POST, request.FILES)
            if form1.is_valid():
                csv_file = request.FILES['file']
                decoded_file = csv_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                csv_reader = csv.reader(io_string, delimiter=',', quotechar='"')

                # Clear all existing records for DataRecord
                DataRecord.objects.all().delete()

                # Collect data for bulk creation
                records = [
                    DataRecord(
                        second_parent_region=row[0],
                        second_parent_id=row[1],
                        rso_id=row[2],
                        rso_msisdn=row[3],
                        retailer_id=row[4],
                        retailer_msisdn=row[5]
                    )
                    for row in csv_reader
                ]

                DataRecord.objects.bulk_create(records)
                return HttpResponse('CSV file for DataRecord uploaded and data inserted successfully.')

        elif 'upload_bvs_xlsx' in request.POST:
            form2 = BVSUploadForm(request.POST, request.FILES)
            if form2.is_valid():
                excel_file = request.FILES['file']
                wb = load_workbook(excel_file)
                sheet = wb.active

                # Clear all existing records for DataRecordBVS
                DataRecordBVS.objects.all().delete()

                records = [
                    DataRecordBVS(Device_ID=row[0], Franchise_ID=row[3])
                    for row in sheet.iter_rows(min_row=3, values_only=True)
                ]

                DataRecordBVS.objects.bulk_create(records)
                return HttpResponse('XLSX file for DataRecordBVS uploaded and data inserted successfully.')

        elif 'upload_heirarchy_xlsx' in request.POST:
            form3 = HeirarchyUploadForm(request.POST, request.FILES)
            if form3.is_valid():
                excel_file = request.FILES['file']
                wb = load_workbook(excel_file)
                sheet = wb["HQ"]  # Explicitly select the "HQ" sheet

                # Clear all existing records for Heirarchy
                Heirarchy.objects.all().delete()

                records = [
                    Heirarchy(Franchise_ID=row[0], Grid=row[1],Cluster=row[2])
                    for row in sheet.iter_rows(min_row=3, values_only=True)
                ]

                Heirarchy.objects.bulk_create(records)
                return HttpResponse('XLSX file for Heirarchy uploaded and data inserted successfully.')

    return render(request, 'upload.html', {'form1': form1, 'form2': form2, 'form3': form3})

from .models import DataRecordBVS
from django.http import JsonResponse
from .models import DataRecordBVS,Heirarchy

from django.http import JsonResponse
from myapp.models import Heirarchy  # Import your model

def get_Cluster(request):
    Franchise_ID = request.GET.get('Franchise_ID')
    if Franchise_ID:
        try:
            # Fetch data from the Heirarchy model
            data_record = Heirarchy.objects.get(Franchise_ID=Franchise_ID)
            print(data_record.Grid)  # Debugging output to verify the value
            return JsonResponse({'Grid': data_record.Grid, 'Cluster': data_record.Cluster})
        except Heirarchy.DoesNotExist:  # Handle the exception properly
            # Return None if the record doesn't exist
            return JsonResponse({'Grid': None, 'Cluster': None}, status=404)
    # Return None if no Franchise_ID is provided
    return JsonResponse({'Grid': None, 'Cluster': None}, status=400)


def get_bvs_number(request):
    retailer_id = request.GET.get('retailer_id')
    if retailer_id:
        try:
            # Fetch data from the DataRecordBVS model
            data_record = DataRecordBVS.objects.get(retailer_id=retailer_id)
            print(data_record.Device_ID)
            return JsonResponse({'Device_ID': data_record.Device_ID})
        except DataRecordBVS.DoesNotExist:
            # Return None if the record doesn't exist
            return JsonResponse({'Device_ID': None})
    # Return None if no retailer_id is provided
    return JsonResponse({'Device_ID': None})

# In your views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Contact

from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse

from django.shortcuts import render, get_object_or_404, redirect
from .models import Contact
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
import json

from django.utils.timezone import now
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Contact,delete_contact  # Ensure this imports your Contact model

@login_required
def contact_delete(request, contact_id):  
    print("Delete view called with ID:", contact_id)  # Debugging

    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        print("Request method is POST")  # Debugging
        user = request.user  # Get the logged-in user

        # Get Pakistan timezone
        pakistan_tz = pytz.timezone('Asia/Karachi')
        pakistan_time = now().astimezone(pakistan_tz)

        # Extract resignation date from request
        resignation_date = request.POST.get('Date_of_Resignation')

        if not resignation_date:  # Ensure it's provided
            return JsonResponse({"error": "Date of Resignation is required"}, status=400)

        # Move record to delete_contact table
        delete_contact.objects.create(
            Retailer_ID=contact.Retailer_ID,
            Franchise_ID=contact.Franchise_ID,
            Retailer_Number=contact.Retailer_Number,
            DSO_Name=contact.DSO_Name,
            CNIC=contact.CNIC,
            BVS_Device=contact.BVS_Device,
            Location=contact.Location,
            username=user.username,  # Store the username of the logged-in user
            Category=contact.Category,
            Other_Contact_Number=contact.Other_Contact_Number,
            Date_of_Joining=contact.Date_of_Joining,
            Date_of_Resignation=resignation_date,  # Use received date
            Date_of_Data_Entry=pakistan_time
        )

        # Delete the original contact
        contact.delete()

        return JsonResponse({"message": "Contact moved to delete_contact and deleted from myapp_contact"}, status=200)

    print("Invalid request method:", request.method)  # Debugging
    return JsonResponse({"error": "Invalid request"}, status=400)


from django.shortcuts import render
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from .models import Contact  # Import your contact model

def generate_oath_pdf(request, contact_id):
    # Retrieve the contact from the database using the contact ID
    contact = Contact.objects.get(id=contact_id)
    
    # Create the PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{contact.DSO_Name}_Oath_Form.pdf"'
    
    # Create a canvas object to generate the PDF
    pdf = canvas.Canvas(response, pagesize=letter)
    
    # Leave half the page blank
    pdf.drawString(100, 100, "")
    
    # Set up the oath statement
    pdf.setFont("Helvetica", 12)
    pdf.drawString(80, 250, "I, _____, S/O _____, work at Zong Franchise __________________ as a ________________.")
    pdf.drawString(80, 230, f"The BVS device assigned to me has the IMEI number {contact.BVS_Device}.")
    pdf.drawString(80, 210, "I understand that I will be held responsible for any unlawful activity involving this device,")
    pdf.drawString(80, 190, "and the company reserves the right to take legal action against me.")
    
    # Add signature line
    pdf.drawString(80, 150, "Signature: ____________________________")
    pdf.drawString(80, 130, "Date: ________________________________")
    
    # Save the PDF
    pdf.showPage()
    pdf.save()
    
    return response


from django.utils.timezone import activate
import pytz

activate(pytz.timezone('Asia/Karachi'))  # Force Django to use Asia/Karachi

